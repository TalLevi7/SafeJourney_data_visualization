import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Function to create an Excel template for the financial data reports

def create_financial_template(output_file='financial_template.xlsx'):
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial_Reports"

    # Define headers
    headers_pl = [
        'Month',
        'Revenue',
        'Cost of Sales',
        'Gross Profit',  # Calculated
        'General expenses',
        'Sales and Marketing Expenses',
        'Salary Expenses',
        'Operating Profit',  # Calculated
        'Financing Expenses',
        'Depreciation  Expenses',
        'Profit Before Tax'  # Calculated
    ]

    headers_cf = [
        'Month',
        'Starting Cash',
        'Profit Before Tax',  # Linked from above
        'Depreciation',  # Linked from above
        'Investments',
        'Loans',
        'Owner financing',
        'Change in Cash',  # Calculated
        'Ending Cash'  # Calculated
    ]

    # Define periods
    periods = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12',
               'Year 1', 'Year 2', 'Year 3']

    # Write Profit & Loss Table

    ws['A1'] = "Profit & Loss Statement"
    ws['A1'].font = Font(bold=True)


    # Write headers
    for row, header in enumerate(headers_pl, 2):
        cell = ws.cell(row=row, column=1)
        cell.value = header
        cell.font = Font(bold=True)
        # cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Write periods
    for col, period in enumerate(periods, 2):
        ws.cell(row=2, column=col, value=period)

    # Add formulas for Profit & Loss
    for col in range(2, 17):  # Columns B to P (2 to 15 in 1-based index)
        col_letter = get_column_letter(col)  # Convert column index to letter

        # Gross Profit = Revenue - Cost of Sales
        ws[f'{col_letter}5'] = f'={col_letter}3-{col_letter}4'

        # Total Operating Expenses
        ws[f'{col_letter}9'] = f'={col_letter}5-{col_letter}6-{col_letter}7-{col_letter}8'

        # Profit Before Tax
        ws[f'{col_letter}12'] = f'={col_letter}9-{col_letter}10-{col_letter}11'


    # Write Cash Flow section Table

    ws['A14'] = "Cash Flow Statement"
    ws['A14'].font = Font(bold=True)

    # Write headers
    for row, header in enumerate(headers_cf, 15):
        cell = ws.cell(row=row, column=1)
        cell.value = header
        cell.font = Font(bold=True)
        # cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Write periods
    for col, period in enumerate(periods, 2):
        ws.cell(row=15, column=col, value=period)


    # Add formulas for Cash Flow
    # Iterate over columns B to P (2 to 16 in Excel column index)
    for col in range(2, 17):
        col_letter = get_column_letter(col)  # Convert column number to letter

        # Row 17 = Row 12
        ws[f'{col_letter}17'] = f'={col_letter}12'

        # Row 18 = Row 11
        ws[f'{col_letter}18'] = f'={col_letter}11'

        # Row 22 = Row 17 + Row 18
        ws[f'{col_letter}22'] = f'={col_letter}17+{col_letter}18'

        # Row 23 = Row 16 + Row 22
        ws[f'{col_letter}23'] = f'={col_letter}16+{col_letter}22'

    # Iterate over columns C to P (3 to 16 in Excel column index)
    for col in range(3, 17):
        col_letter = get_column_letter(col)  # Current column
        prev_col_letter = get_column_letter(col - 1)  # Previous column

        # Row 16 in the current column = Row 23 in the previous column
        ws[f'{col_letter}16'] = f'={prev_col_letter}23'
        if col_letter == 'N':
            ws['N16'] = '=B16'


    # Save the workbook
    wb.save(output_file)

    print("Excel template has been created")
